##coding=utf-8
#教务爬虫

#导包
import requests
import re
import time
import pytesseract
from PIL import Image,ImageEnhance,ImageFilter
import json
import operator
from bs4 import BeautifulSoup
from openpyxl import Workbook
from prettytable import PrettyTable
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
import os.path
#发送邮件的邮箱
my_sender ="****@163.com"
password = "****"
sender = my_sender
receivers = ','.join(["****@qq.com"])
smtp_server = "smtp.163.com"

global termCode

#requests 对象重定义   保持cookies登录状态
requests = requests.Session()


def Choice_Term():
    print("--------菜单列表--------")
    print("1、2016年夏季学期")
    print("2、2016年秋季学期")
    print("3、2017年春季学期")
    print("4、2017年夏、秋季学期")
    print("5、2018年春季学期")
    print("6、其它")
    i = input("请选择学期：")
    global termCode

    #选择学期
    if i=='1':
        termCode = "2016-2017-0"
    elif i=='2':
        termCode = "2016-2017-1"
    elif i=='3':
        termCode = "2016-2017-2"
    elif i=='4':
        termCode = "2017-2018-1"
    elif i=='5':
        termCode = "2017-2018-2"
    elif i=='6':
        print("学年+学期（例如：2015到2016年的第一个学期 2015-2016-1）")
        termCode = input("请输入学期：")
    else:
        print("输入有误！请重新选择")
        Choice_Term()
        
class JW(object):

    global termCode
    def __init__(self,id="",password=""):
        self._id = id
        self._password = password
        #网址  头header
        self._headers = {
           "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.79 Safari/537.36 Edge/14.14393"
           ,"Referer": "http://jwzx.usc.edu.cn/Login/Login"
           }
        self.login()

    
    #登录
    def login(self):
        #得到验证码图片
        response = requests.get("http://jwzx.usc.edu.cn/Core/verify_code.ashx",headers = self._headers)
        codeImage = response.content
        fn = open("code.png","wb")
        fn.write(codeImage)
        fn.close()
        global FLAG
        FLAG = "1"
        #数据
        data = {
            "UserName": self._id,
            "Password": self._password,
            "Code":self.know_V_code()
            }
        #登录请求
        response = requests.post("http://jwzx.usc.edu.cn/Login/Login",data=data,headers = self._headers)
        #判断登录是否成功
        if '1' in response.text:
            print("登录成功！")
            return True
        #失败则循环login()方法 直至成功
        else:
            #打印登录失败原因
            print(response.text)
            if "验证码" in response.text:
                self.login()
            else:
                exit()

    #验证码识别
    def know_V_code(self):
        image = Image.open('code.png')
        
        #二值化
        #enhancer = ImageEnhance.Contrast(image)
        #image = enhancer.enhance(2)
        #image = image.convert('1')
        #image.show()
        #data = image.getdata()
        
        #w,h = image.size()
        #vcode = input("请输入验证码：")
        
        vcode = pytesseract.image_to_string(image)
        return vcode

    #菜单
    def Menu(self):
        
        
        while True:
            print("--------菜单列表--------")
            print("1、获取成绩")
            print("2、获取课表")
            print("3、一键评教")
            print("4、获取考试安排")
            print("5、重新选择学期")
            print("6、退出")
            i = input("请输入")
            if i=='1':
                self.GetScore()
                
            elif i=='2':
                self.GetStudentTimetable()
                
            elif i=='3':
                self.StuEvaluation()
                
            elif i=='4':
                self.GetExamSitelist()
                
            elif i=='5':
                Choice_Term()
                a.Menu()
            elif i=='6':
                response = requests.post("http://jwzx.usc.edu.cn/Login/LogOut",data="",headers = self._headers)
                print(response.text+"\n注销成功！")
                exit()
            else:
                print("输入错误!\n")
            print("have running")
            
        self.Menu()

    #评教
    def StuEvaluation(self):
        #得到评教批次 (学期)
        data = {
            "order":"ASC",
            "sort":"BatchName"
            }
        response = requests.post("http://jwzx.usc.edu.cn/Student/StuEvaluation/GetBatchList",data=data,headers = self._headers)
        data = str(response.content,encoding='utf-8')
        result = json.loads(data)
        #Id有效
        Id="" 
        #选择学期
        Str=termCode
        for k in result['rows']:
            #查找Id
            if Str==k['TermCode']:
                Id=k['Id']
            
            

        #得到该学期下的课程列表
        data = {
            "sort": "CourseName",
            "order": "ASC",
            "batchId": Id
            }
        response = requests.post("http://jwzx.usc.edu.cn/Student/StuEvaluation/GetCourseList",data=data,headers = self._headers)
        data = str(response.content,encoding='utf-8')
        result = json.loads(data)
        for i in result['rows']:
            #执行各门课程的评教(根据请求Id的不同)
            Id=i['Id']
            data = {
                "sort": "SerialNo",
                "order": "ASC",
                "evaluationValuatorId": Id
                }
            response = requests.post("http://jwzx.usc.edu.cn/Student/StuEvaluation/GetResultList",data=data,headers = self._headers)
            data = str(response.content,encoding='utf-8')
            result = json.loads(data)
            count=0
            for j in result['rows']:
                count=count+1
                #执行评教选项
                Id_1=j['Id']
                if count==1:
                    grade=3
                else:
                    grade=4
                data = {
                    "rankCode": grade,
                    "resultId": Id_1
                    }
                response = requests.post("http://jwzx.usc.edu.cn/Student/StuEvaluation/UpdateEvaluationResult",data=data,headers = self._headers)
                if '1' in response.text:
                    continue
                else:
                    print("第%d个评教选项失败"%count)
            data={
                "evaluationValuatorId":Id
                }
            response = requests.post("http://jwzx.usc.edu.cn/Student/StuEvaluation/SubmitEvaluationResult",data=data,headers = self._headers)
            if '1' in response.text:
                continue
            else:
                print(response.text)

    
    #成绩 (剩：表格界面优化)
    def GetScore(self):
        data = {
            "termCode":termCode,
            "sort":"Id",
            "order":"ASC"
            }
        #文件操作
        workbook=Workbook()
        #读取成绩
        response = requests.post("http://jwzx.usc.edu.cn/Student/StuTermCourseScore/GetList",data=data,headers = self._headers)
        
        data = str(response.content,encoding='utf-8')
        global FLAG
        #将字符串转换为字典
        result = json.loads(data)
        #得到成绩门数
        length = result['total']
        #将成绩输出到Excel表格中
        title = ["类别","学期","班级名称","学号","姓名","性别","课程代码","课程名称","学分","考核","类型","教学班级","平时成绩","考试成绩","总评成绩","绩点"]
        #创建表
        ws = workbook.active
        a='A'
        i=1
        
        Grode = ""

        #打印title
        for j in title:
            ws[a+str(i)]=j
            
            Grode +=j+'\t'
            a=chr(ord(a)+i)
        a='A'
        row=2
        Grode +='\n\n'
        #打印成绩
        for i in result['rows']:
            for k,v in i.items():
                if v!=None:
                    ws[a+str(row)]=v

                    Grode +=str(v)+'\t'
                    a=chr(ord(a)+1)
            a='A'
            Grode +='\n\n'
            row=row+1
        #保存
        workbook.save('成绩.xlsx')
        print("成绩打印成功！")
        
        if operator.eq(data,FLAG)!=True:

            FLAG = data
            msg = MIMEMultipart()
            msg['Subject'] = '成绩呀呀呀呀'
            msg['From'] = sender
            msg['To'] = receivers
            xlsxpart = MIMEText(open("成绩.xlsx",'rb').read(),'base64','utf-8')
            xlsxpart["Content-Type"]='application/octet-stream'
            xlsxpart["Content-Disposition"] = 'attachment;filename="111.xlsx"'.format(xlsxpart)
            #编码
            msg.attach(xlsxpart)
            message = "成绩\n\n\n"+Grode
            body = MIMEText(message,"utf-8")
            msg.attach(body)
            clicent = smtplib.SMTP_SSL()
            clicent.connect(smtp_server,465)
            clicent.set_debuglevel(1)
            clicent.helo(smtp_server)
            clicent.ehlo(smtp_server)
            clicent.login(my_sender,password)
            try:
                clicent.sendmail(sender,receivers,msg.as_string())
                clicent.close()
                print("发送成功！")
            except smtplib.SMTPRecipientsRefused:
                print("Recipient refused")
            except smtplib.SMTPAuthenticationError:
                print("Auth error")
            except smtplib.SMTPSenderRefused:
                print("Sender refused")
            except smtplib.SMTPException as e:
                print(e)
            

    #考试
    def GetExamSitelist(self):
        data = {
            "termCode":termCode,
            "sort":"CourseName",
            "order":"ASC"
            }
        #文件操作
        workbook=Workbook()
        #获取考试安排请求
        response = requests.post("http://jwzx.usc.edu.cn/Student/StuExamSite/GetExamSitelist",data = data,headers = self._headers)
        data = str(response.content,encoding='utf-8')
        #将字符串转换为字典
        result = json.loads(data)
        #得到考试门数
        length = result['total']
        
        title = ["考场编号","课程代码","课程名称","考试日期","考试星期","考试时间","考试地点","校区"]
        ws = workbook.active
        a='A'
        i=1
        #打印title
        for j in title:
            ws[a+str(i)]=j
            a=chr(ord(a)+i)
        a='A'
        row=2
        #打印成绩
        for i in result['rows']:
            for k,v in i.items():
                if v!=None:
                    ws[a+str(row)]=v
                    a=chr(ord(a)+1)
            a='A'
            row=row+1
        #保存
        workbook.save('考试安排.xlsx')
        print("考试安排打印成功！")


    #课表  (剩：表格界面优化)
    def GetStudentTimetable(self):
        data = {
            "sort": "TimeName",
            "order": "ASC",
            "termCode": termCode
            }
        #文件操作
        workbook=Workbook()
        #读取课表请求
        response = requests.post("http://jwzx.usc.edu.cn/Student/StuTimetable/GetStudentTimetable",data=data,headers = self._headers)
   
        response.encoding = response.apparent_encoding
        
        data = str(response.content,encoding='utf-8')
        
        
        tb = PrettyTable()
        #json转换字典
        result = json.loads(data)
        
        #得到总课程数
        length = result['total']
        #将成绩输出到Excel表格中
        title = ["编号","节次","星期天","星期一","星期二","星期三","星期四","星期五","星期六"]
        tb.field_names =title
        #创建表
        ws = workbook.active
        a='A'
        i=1
        #打印title
        for j in title:
            ws[a+str(i)]=j
            #垂直和水平居中
            tb.align[j]="c"
            tb.valign[j]="t"
            a=chr(ord(a)+i)
        a='A'
        row=2

        #json必备
        headStr = '{\n"classInfo":[\n'
        tailStr = ']\n}'
        itemHeadStr = '{\n'
        itemTailStr = '\n}'
        classInfoStr = ''
        classInfoStr +=headStr
        count = 0
        
        #打印课表
        for i in result['rows']:
            arr = []
            
            classInfoStr +='['
            for k,v in i.items():
                eee=str(v)
                
                if eee[:1]!='<':
                    
                    ws[a+str(row)]=v
                    arr.append(v)
                    
                else:
                    soup = BeautifulSoup(str(v))
                    cls = soup.find_all("li")
                    Str =""
                    itemClassInfoStr = ""
                    
                    #判断同一节次的课程数目
                    Len = 0
                    #若同一节次有着不同时间的2门课程
                    #一节课有5个li标签 Len每次加5
                    while Len <len(cls):
                        itemClassInfoStr  = itemHeadStr + '"SUMMARY":"' + cls[0+Len].string+'-'+cls[2+Len].string+'|'+arr[1]+'|'+cls[4+Len].string+ '",\n'
                        itemClassInfoStr +='"weekday":"'+str(len(arr)-2)+ '",\n'
                        itemClassInfoStr +='"week":"'+cls[3+Len].string[:-5]+ '"\n'
                        itemClassInfoStr += itemTailStr
                        classInfoStr += itemClassInfoStr+','
                        Len +=5
                    
                    #Str=cls[0].string+'-'+cls[2].string+'|'+title[len(arr)]+arr[1]+'|'+cls[4].string+'|'+cls[3].string
                    #print(Str)
                    #print()
                    for j in cls:
                        Str=Str+j.string
                    ws[a+str(row)]=Str
                    arr.append(Str)
                a=chr(ord(a)+1)
            a='A'
            classInfoStr +=']'
            if count!=len(result['rows'])-1:
                classInfoStr +=','
            count +=1
            tb.add_row(arr)
            row=row+1
        classInfoStr +=tailStr
        #print(tb)
        #保存
        workbook.save('课表.xlsx')
        print("课表打印成功！")
        #写入json
        with open('classInfo.json','w') as f:
            #
            #写入json文件前一定得先调用dumps方法进行系列化
            #否则在调用loads方法反序列化时会报错（json.decoder.JSONDecodeError）
            #
            f.write(json.dumps(classInfoStr))
            f.close()
            





classNum = input("请输入学号：")
pw = input("请输入密码：")
a = JW(classNum,pw)
Choice_Term()
a.Menu()
        
